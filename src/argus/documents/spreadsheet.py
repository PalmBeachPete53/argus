from __future__ import annotations

import csv
import io
import re
from pathlib import Path

from ..normalize import title_from_url
from .base import (
    METHOD_CSV,
    METHOD_XLSX,
    WARNING_EMPTY_TEXT,
    DocumentParser,
    DocumentTable,
    NormalizedDocument,
)
from ._util import decode_text, make_unavailable


def _matrix_rows(rows) -> list[list[str]]:
    """Trim trailing empty rows and trailing empty columns; stringify cells."""
    matrix: list[list[str]] = []
    for row in rows:
        if row is None:
            matrix.append([])
            continue
        matrix.append(["" if cell is None else str(cell).strip() for cell in row])
    while matrix and not any(cell for cell in matrix[-1]):
        matrix.pop()
    if not matrix:
        return matrix
    width = max(len(row) for row in matrix)
    trimmed = []
    for row in matrix:
        row = row + [""] * (width - len(row))
        while row and not row[-1]:
            row.pop()
        trimmed.append(row)
    return trimmed


def _matrix_to_table(order: int, name: str, matrix: list[list[str]], **metadata) -> DocumentTable:
    headers = matrix[0] if matrix else []
    body = matrix[1:] if len(matrix) > 1 else []
    return DocumentTable(order=order, name=name, headers=headers, rows=body, metadata=metadata)


class XlsxParser(DocumentParser):
    kind = "xlsx"
    label = "Generic XLSX parser (openpyxl, worksheet -> structured tables)"
    extraction_method = METHOD_XLSX

    def parse(self, document) -> NormalizedDocument:
        if document.local_path is None:
            return make_unavailable(document)
        try:
            import openpyxl

            workbook = openpyxl.load_workbook(document.local_path, read_only=True, data_only=True)
        except Exception as exc:
            doc = make_unavailable(document, warnings=["parse_error"])
            doc.metadata["parse_error"] = f"{exc.__class__.__name__}: {exc}"
            return doc

        tables: list[DocumentTable] = []
        sheet_names: list[str] = []
        for sheet_index, sheet_name in enumerate(workbook.sheetnames):
            try:
                sheet = workbook[sheet_name]
                rows = sheet.iter_rows(values_only=True)
                matrix = _matrix_rows(rows)
            except Exception as exc:
                doc = make_unavailable(document, warnings=["parse_error"])
                doc.metadata["parse_error"] = f"{exc.__class__.__name__}: {exc}"
                return doc
            sheet_names.append(sheet_name)
            if not matrix:
                continue
            tables.append(
                _matrix_to_table(
                    sheet_index,
                    sheet_name,
                    matrix,
                    sheet_index=sheet_index,
                    sheet=sheet_name,
                )
            )

        title = None
        try:
            props = workbook.properties
            if props is not None and props.title:
                title = props.title
        except Exception:
            pass

        metadata: dict = {
            "sheet_count": len(sheet_names),
            "sheet_names": sheet_names,
        }
        text = "\n\n".join(table.render() for table in tables)
        warnings = []
        if not text:
            warnings.append(WARNING_EMPTY_TEXT)

        return NormalizedDocument(
            publication_id=document.publication_id,
            document_id=document.sha256 or "",
            source_url=document.url,
            local_path=document.local_path,
            document_kind=document.kind,
            mime_type=document.content_type,
            title=title,
            text=text,
            sections=[],
            tables=tables,
            metadata=metadata,
            extraction_method=self.extraction_method,
            extraction_warnings=warnings,
        )


class CsvParser(DocumentParser):
    kind = "csv"
    label = "Generic CSV parser (stdlib csv, dialect-sniffed)"
    extraction_method = METHOD_CSV

    def parse(self, document) -> NormalizedDocument:
        if document.local_path is None:
            return make_unavailable(document)
        try:
            data = Path(document.local_path).read_bytes()
        except OSError:
            return make_unavailable(document)

        text_value, _ = decode_text(data)
        sample = text_value[:4096]
        delimiter = ","
        try:
            detected = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            delimiter = detected.delimiter
        except Exception:
            delimiter = ","
        try:
            reader = csv.reader(io.StringIO(text_value), delimiter=delimiter)
            matrix = _matrix_rows(reader)
        except Exception as exc:
            doc = make_unavailable(document, warnings=["parse_error"])
            doc.metadata["parse_error"] = f"{exc.__class__.__name__}: {exc}"
            return doc

        if not matrix:
            return make_unavailable(document, warnings=[WARNING_EMPTY_TEXT])

        name = re.sub(r"\.[^.]+$", "", Path(document.local_path).name) or "data"
        table = _matrix_to_table(0, name, matrix)
        metadata: dict = {"delimiter": delimiter, "table_count": 1}
        title = title_from_url(document.url, fallback=name)

        return NormalizedDocument(
            publication_id=document.publication_id,
            document_id=document.sha256 or "",
            source_url=document.url,
            local_path=document.local_path,
            document_kind=document.kind,
            mime_type=document.content_type,
            title=title,
            text=table.render(),
            sections=[],
            tables=[table],
            metadata=metadata,
            extraction_method=self.extraction_method,
            extraction_warnings=[],
        )